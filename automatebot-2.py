from openpyxl import load_workbook  
import pyautogui as pg
import pyperclip
import random
import time 
import wikipedia
import webbrowser
import os

url = 'http://uncle-machine.com/'
webbrowser.open(url) #ให้มันเปิดเวบอัตโนมัติ
time.sleep(5) # ตั้งหน่วงเวลา 3 วิ ให้เน็ททำงาน

#path = r'C:\Users\perzyploy\Documents\python\Bootcamp\EP5-6' #ใส่path ที่เก็บรูปภาพ
#bt1 = os.path.join(path,'product.png') # เรียก path มาใช้

pg.click(r'C:\Users\perzyploy\Documents\python\Bootcamp\EP5-6\product.png')
time.sleep(2)

#pg.moveTo(100,200)


excelfile = load_workbook('fruit.xlsx')
sheet = excelfile.active
#print(sheet['B1'].value)

count = len(sheet['B']) #นับคอลัมน์ B ในไฟล์ exel
#print(count)


fruit_list = []

for i in range(2,count+1): #เริ่มนับที่ row ที่ 2 เพราะไม่เอาหัวข้อ
	data = sheet.cell(row=i, column=2).value
	print(data)
	split = data.split(',') # เป็นคำสั่งให้แยกข้อความเปป็นสองส่วน โดยใช้คอมมาเป็นตัวกรอง จาก index เดียวเป็นสอง index
	if len(split) >= 2: # สร้างเงื่อนไขว่าถ้าจำนวน obj ที่ถูกสปลิทมากกว่าหรือเท่ากับ 2 ให้ปรินท์ split ออกมา
		#print(split[0]) # ใส่เงื่อนไขว่าให้ปรินท์เฉพาะ index 0 เท่านั้น เพื่อตัดข้อความที่ไม่ใช้ออก
		if split[0] not in fruit_list: # สร้างเงือนไขว่าถ้า split index 0 ไม่อยู่ใน fruit list ให้ทำการเพิ่มเข้าไป
			fruit_list.append(split[0]) # เอาเฉพาะ split index 0 ไปใส่เพ่ิมใน list ที่ชื่อ fruit_list

	else:
		split = data.split('(') # ถ้าจำนวน obj หลังสปลิทไม่มากกว่าหรือเท่ากับ 2 ให้ปปลิทอีกครั้งโดยใช้วงเล็บเปิดเป็นตัวกรอง
		print(split)
		if split[0] not in fruit_list: # สร้างเงือนไขว่าถ้า split index 0 ไม่อยู่ใน fruit list ให้ทำการเพิ่มเข้าไป
			fruit_list.append(split[0])
		

	fruit_list.append(data)

print(fruit_list)

# Work Flow

# คลิกที่เว็บบราวเซอร์
pg.click(100,200) # สั่งให้เมาส์คลิกหน้าจอที่พิกัด 100x200


for f in fruit_list:
	# กด tab 6 ครั้งเพื่อให้ cursor อยู่ในช่องกรอก
	pg.press('tab',presses=6) # สั่งให้กด tab ทั้งหมด 6 ครั้ง

	# กรอกข้อมูลชื่อ + tab
	product = f
	pyperclip.copy(product) #เพื่อจะ control v
	pg.hotkey('ctrl','v')
	pg.press('tab')

	# สุ่มราคา
	rand = random.choice(range(100,1001,100))
	# ให้มันสุ่มราคาจาก 100 ถึง index 1001 ซึ่งก็จะสุดที่ 1000 ขยับ step ละ 100 สามารถดู module random ได้จากเน็ท


	# กรอกราคา (ตัวเลข) + tab
	pyperclip.copy(rand)
	pg.hotkey('ctrl','v')
	pg.press('tab')

	try:
		info = wikipedia.summary(f)
	except:
		info = 'No detail-Candy fruit shop'

	# ข้อมูลสินค้า กรอกชื่อร้านไปก่อน+ ะฟิ
	pyperclip.copy(info)
	pg.hotkey('ctrl','v')
	pg.press('tab')


	# tab เพื่อไปยังปุ่ม submit
	pg.press('tab')

	# enter เพื่อ Submit
	pg.press('enter')

time.sleep(1)
